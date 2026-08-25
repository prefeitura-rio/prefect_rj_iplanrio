"""
Compliance Validator Module

This module contains the main compliance validation logic:
- compute_classification(): Determines if an NF is "OK", "Suspect", or "Not Analyzable"
- ComplianceValidator: Main validator class for validating extracted NFs against expected NFs
- validate_against_expected(): Convenience function for quick validation

``ComplianceValidator`` composes plain-function modules (``core``, ``matching``,
``validate``, ``report``) rather than mixin classes: each module operates on an
explicit ``validator`` instance passed as its first argument, and the class
below just holds state and delegates.
"""

from pathlib import Path

from iplanrio_agent_toolkit.rules import Rule, RuleEngine

from . import core, report, validate
from .core import _LEGACY_FALLBACK_RESULT
from .rules import DEFAULT_RULES
from .validation_context import ValidationContext


def compute_classification(
    nf_found: bool,
    valor_pago: float | None = None,
    valor_documento: float | None = None,
    valor_extracted: float | None = None,
    tipo_documento: str | None = None,
    page_categories: list[str] | None = None,
    date_valid: bool | None = None,
    is_duplicate: bool | None = None,
) -> str:
    """
    [DEPRECATED] Use ``core.classify`` or ``RuleEngine.evaluate`` instead.

    This function is kept for backward compatibility and now delegates to the rule engine.

    Compute classification: "OK", "Suspect", or "Not Analyzable"

    :param nf_found: Whether NF was found by extractor.
    :param valor_pago: Amount paid (from database).
    :param valor_documento: Expected total value (from database).
    :param valor_extracted: Extracted total value (from model).
    :param tipo_documento: Document type extracted by model.
    :param page_categories: List of page categories from classifier (optional).
    :param date_valid: Whether NF issue date is valid against company start date (optional).
    :param is_duplicate: Whether this NF appears in multiple different PDFs (optional).
    :returns: Classification: "OK", "Suspect", or "Not Analyzable".
    """
    # Create temporary rule engine with default rules
    engine = RuleEngine(DEFAULT_RULES)

    # Build context
    context = ValidationContext(
        nf_found=nf_found,
        valor_pago=valor_pago,
        valor_documento=valor_documento,
        valor_extracted=valor_extracted,
        tipo_documento=tipo_documento,
        page_categories=page_categories,
        date_valid=date_valid,
        is_duplicate=is_duplicate,
    )

    # Evaluate rules
    result = engine.evaluate(context, fallback=_LEGACY_FALLBACK_RESULT)

    return result.classification


class ComplianceValidator:
    """
    Validates extraction results against expected NFs.

    This is designed to be integrated as the final step in the extraction pipeline.
    """

    def __init__(
        self,
        expected_nfs: list[dict] = None,
        use_bigquery_deduplication: bool = True,
        service_account_path: Path | None = None,
        rules: list[Rule[ValidationContext]] | None = None,
        min_match_score: int = 2,
    ):
        """
        Initialize validator with expected NFs to search for.

        :param expected_nfs: List of expected NF dicts with keys:
            - pdf_name: Name of PDF file
            - cnpj: Expected CNPJ (any format)
            - numero_nf: Expected NF number (any format)
            - valor_total: Expected total value (any format)
            - page: Expected page number (optional, for reporting)
        :param use_bigquery_deduplication: If True, query BigQuery for
            deduplication lookup. If False, use only expected_nfs (for testing).
        :param service_account_path: Optional path to BigQuery credentials.
        :param rules: Optional custom rule list (defaults to DEFAULT_RULES).
        :param min_match_score: Minimum number of fields (CNPJ + número +
            data_emissão) that must match for a declaration to be considered
            found (2 = legacy 2/3 fallback, 3 = strict perfect match only).
            Default: 2.
        """
        core.initialize(
            self,
            expected_nfs=expected_nfs,
            use_bigquery_deduplication=use_bigquery_deduplication,
            service_account_path=service_account_path,
            rules=rules,
            min_match_score=min_match_score,
        )

    def validate_extraction(
        self,
        pdf_name: str,
        extracted_nfs: list[dict],
        page_categories: list[str] | None = None,
    ) -> dict:
        """Validate extraction results for a single PDF against expected NFs. See ``validate.validate_extraction``."""
        return validate.validate_extraction(self, pdf_name, extracted_nfs, page_categories)

    def validate_batch(self, extraction_results: list[dict]) -> dict:
        """Validate a batch of extraction results. See ``validate.validate_batch``."""
        return validate.validate_batch(self, extraction_results)

    def print_validation_report(self, validation: dict, verbose: bool = True) -> None:
        """Log a human-readable validation report for a single PDF. See ``report.print_validation_report``."""
        report.print_validation_report(validation, verbose)

    def print_batch_report(self, batch_validation: dict, group_by_nf: bool = True) -> None:
        """Log aggregate batch validation report. See ``report.print_batch_report``."""
        report.print_batch_report(batch_validation, group_by_nf)

    @staticmethod
    def load_expected_nfs_from_excel(excel_path: Path) -> list[dict]:
        """
        Load expected NFs from validation Excel file.

        :param excel_path: Path to validation Excel file with NF_Details sheet.
        :returns: List of expected NF dicts.
        :raises KeyError: If the sheet ``NF_Details`` is not found in the workbook.
        :raises ValueError: If required columns are missing from the sheet header.
        """
        from openpyxl import load_workbook

        wb = load_workbook(excel_path, read_only=True, data_only=True)
        if "NF_Details" not in wb.sheetnames:
            raise KeyError(f"Sheet 'NF_Details' not found in '{excel_path}'. Available sheets: {wb.sheetnames}")
        ws = wb["NF_Details"]

        rows = ws.iter_rows(values_only=True)
        header = next(rows)
        col_idx = {name: i for i, name in enumerate(header) if name is not None}

        required_cols = {"PDF_name", "CNPJ", "Numero_NF", "Valor_Total"}
        missing = required_cols - col_idx.keys()
        if missing:
            raise ValueError(f"Required columns missing from 'NF_Details' sheet: {sorted(missing)}")

        page_col = col_idx.get("NF_Page")
        expected_nfs = []
        for row in rows:
            expected_nfs.append(
                {
                    "pdf_name": row[col_idx["PDF_name"]],
                    "cnpj": row[col_idx["CNPJ"]],
                    "numero_nf": row[col_idx["Numero_NF"]],
                    "valor_total": row[col_idx["Valor_Total"]],
                    "page": row[page_col] if page_col is not None else "Unknown",
                }
            )

        wb.close()
        return expected_nfs


# Convenience function for quick validation
def validate_against_expected(
    pdf_name: str,
    extracted_nfs: list[dict],
    expected_nfs: list[dict],
    page_categories: list[str] | None = None,
) -> dict:
    """
    Quick validation of extraction results against expected NFs.

    :param pdf_name: Name of PDF file.
    :param extracted_nfs: List of extracted NF dicts.
    :param expected_nfs: List of expected NF dicts (all PDFs).
    :param page_categories: List of page categories from classifier (optional).
    :returns: Validation result dict with classification
        (OK/Suspect/Not Analyzable).
    """
    validator = ComplianceValidator(expected_nfs)
    return validator.validate_extraction(pdf_name, extracted_nfs, page_categories)
