"""
Compliance Validator Module

This module contains the main compliance validation logic:
- compute_classification(): Determines if an NF is "OK", "Suspect", or "Not Analyzable"
- ComplianceValidator: Main validator class for validating extracted NFs against expected NFs
- validate_against_expected(): Convenience function for quick validation
"""

import logging
from pathlib import Path

logger = logging.getLogger(__name__)

# Import utility functions from the same package
# Import rule engine components
from .rule_engine import RuleEngine
from .rules import DEFAULT_RULES
from .rules.base import ComplianceRule, RuleResult
from .utils import check_date_against_company_start, normalize_cnpj, normalize_number, normalize_value
from .validation_context import ValidationContext


def compute_classification(
    nf_found: bool,
    valor_pago: float | None = None,
    valor_documento: float | None = None,
    valor_extracted: float | None = None,
    tipo_documento: str | None = None,
    page_categories: list[str] | None = None,
    date_valid: bool | None = None,
    is_duplicate: bool | None = None,
) -> str:
    """
    [DEPRECATED] Use ComplianceValidator._classify or RuleEngine.evaluate instead.

    This function is kept for backward compatibility and now delegates to the rule engine.

    Compute classification: "OK", "Suspect", or "Not Analyzable"

    :param nf_found: Whether NF was found by extractor.
    :param valor_pago: Amount paid (from database).
    :param valor_documento: Expected total value (from database).
    :param valor_extracted: Extracted total value (from model).
    :param tipo_documento: Document type extracted by model.
    :param page_categories: List of page categories from classifier (optional).
    :param date_valid: Whether NF issue date is valid against company start date (optional).
    :param is_duplicate: Whether this NF appears in multiple different PDFs (optional).
    :returns: Classification: "OK", "Suspect", or "Not Analyzable".
    """
    # Create temporary rule engine with default rules
    engine = RuleEngine(DEFAULT_RULES)

    # Build context
    context = ValidationContext(
        nf_found=nf_found,
        valor_pago=valor_pago,
        valor_documento=valor_documento,
        valor_extracted=valor_extracted,
        tipo_documento=tipo_documento,
        page_categories=page_categories,
        date_valid=date_valid,
        is_duplicate=is_duplicate
    )

    # Evaluate rules
    result = engine.evaluate(context)

    return result.classification


class ComplianceValidator:
    """
    Validates extraction results against expected NFs.

    This is designed to be integrated as the final step in the extraction pipeline.
    """

    def __init__(
        self,
        expected_nfs: list[dict] = None,
        use_bigquery_deduplication: bool = True,
        service_account_path: Path | None = None,
        rules: list[ComplianceRule] | None = None,
        min_match_score: int = 2,
    ):
        """
        Initialize validator with expected NFs to search for.

        :param expected_nfs: List of expected NF dicts with keys:
            - pdf_name: Name of PDF file
            - cnpj: Expected CNPJ (any format)
            - numero_nf: Expected NF number (any format)
            - valor_total: Expected total value (any format)
            - page: Expected page number (optional, for reporting)
        :param use_bigquery_deduplication: If True, query BigQuery for
            deduplication lookup. If False, use only expected_nfs (for testing).
        :param service_account_path: Optional path to BigQuery credentials.
        :param rules: Optional custom rule list (defaults to DEFAULT_RULES).
        :param min_match_score: Minimum number of fields (CNPJ + número +
            data_emissão) that must match for a declaration to be considered
            found. 2 = allow 2/3 match with fallback (default, legacy
            behaviour). 3 = require all 3 fields to match (strict / no fallback).
        """
        self.expected_nfs = expected_nfs or []
        self.use_bigquery_deduplication = use_bigquery_deduplication
        self.service_account_path = service_account_path
        self.min_match_score = min_match_score

        # Initialize rule engine
        self.rule_engine = RuleEngine(rules or DEFAULT_RULES)

        # Build normalized lookup for fast matching
        self._build_expected_lookup()

    def _build_expected_lookup(self):
        """Build normalized lookup structure for expected NFs."""
        self.expected_by_pdf = {}
        self.pdf_data_envio = {}  # Maps pdf_name -> data_envio for duplicate detection

        # Build deduplication lookup from BigQuery OR from expected_nfs
        if self.use_bigquery_deduplication:
            # Query BigQuery for complete deduplication lookup
            from run_poc.bigquery_loader import (
                get_deduplication_lookup_from_bigquery,
            )

            try:
                logger.info("Loading deduplication lookup from BigQuery...")
                self.deduplication_lookup = get_deduplication_lookup_from_bigquery(
                    service_account_path=self.service_account_path
                )
                logger.info(
                    "Loaded %d unique (CNPJ, Numero) combinations",
                    len(self.deduplication_lookup),
                )

                # Count duplicates
                duplicates = sum(1 for pdfs in self.deduplication_lookup.values() if len(pdfs) > 1)
                logger.info("Found %d combinations appearing in multiple PDFs", duplicates)
            except Exception as e:
                logger.warning("Failed to load BigQuery deduplication lookup: %s", e)
                logger.warning("Falling back to local expected_nfs for deduplication")
                self.deduplication_lookup = {}
                self.use_bigquery_deduplication = False
        else:
            # Fallback: Build from local expected_nfs only (for testing)
            self.deduplication_lookup = {}

        # Build expected_by_pdf and local deduplication (if BigQuery failed)
        for nf in self.expected_nfs:
            pdf_name = nf['pdf_name']

            if pdf_name not in self.expected_by_pdf:
                self.expected_by_pdf[pdf_name] = []

            # Store data_envio for this PDF (for duplicate detection)
            if 'data_envio' in nf and pdf_name not in self.pdf_data_envio:
                self.pdf_data_envio[pdf_name] = nf['data_envio']

            cnpj_norm = normalize_cnpj(nf['cnpj'])
            numero_norm = normalize_number(nf['numero_nf'])
            valor_norm = normalize_value(nf['valor_total'])

            self.expected_by_pdf[pdf_name].append({
                'original': nf,
                'cnpj_norm': cnpj_norm,
                'numero_norm': numero_norm,
                'valor_norm': valor_norm,
                'cod_organizacao': nf.get('cod_organizacao'),  # For duplicate detection
                'cod_unidade': nf.get('cod_unidade'),  # For duplicate detection
                'id_documento': nf.get('id_documento'),  # For duplicate detection tie-breaking
                'matched': False  # Track if this expected NF was found
            })

            # If not using BigQuery, build local deduplication from expected_nfs
            if not self.use_bigquery_deduplication:
                # 4-field dedup key: exact (cnpj, numero, org, unit) combination must repeat
                cod_org = nf.get('cod_organizacao', '')
                cod_unit = nf.get('cod_unidade', '')
                dedup_key = (cnpj_norm, numero_norm, cod_org, cod_unit)

                if dedup_key not in self.deduplication_lookup:
                    self.deduplication_lookup[dedup_key] = []

                # Store fields matching BigQuery structure
                entry = {
                    'pdf_name': pdf_name,
                    'data_envio': nf.get('data_envio'),
                    'id_documento': nf.get('id_documento'),
                    'cod_organizacao': cod_org,
                    'cod_unidade': cod_unit
                }

                # Only add if not already in list
                if not any(e['pdf_name'] == pdf_name for e in self.deduplication_lookup[dedup_key]):
                    self.deduplication_lookup[dedup_key].append(entry)

    def _classify(
        self,
        nf_found: bool,
        valor_pago: float | None = None,
        valor_documento: float | None = None,
        valor_extracted: float | None = None,
        tipo_documento: str | None = None,
        page_categories: list[str] | None = None,
        date_valid: bool | None = None,
        is_duplicate: bool | None = None,
        data_emissao_expected: str | None = None,
        data_emissao_extracted: str | None = None,
        data_servico: str | None = None,
        cnpj_data_abertura: str | None = None,
    ) -> RuleResult:
        """
        Classify NF using rule engine.

        This replaces the old compute_classification function.

        :param nf_found: Whether NF was found by extractor.
        :param valor_pago: Amount paid (from database).
        :param valor_documento: Expected total value (from database).
        :param valor_extracted: Extracted total value (from model).
        :param tipo_documento: Document type extracted by model.
        :param page_categories: List of page categories from classifier (optional).
        :param date_valid: Whether NF issue date is valid against company start date (optional).
        :param is_duplicate: Whether this NF appears in multiple different PDFs (optional).
        :param data_emissao_expected: Expected emission date from declaration (optional).
        :param data_emissao_extracted: Extracted emission date from PDF (optional).
        :param data_servico: Service provision date from PDF (optional).
        :param cnpj_data_abertura: Company opening date from BigQuery (optional).
        :returns: RuleResult with classification, rule_name, and reason.
        """
        # Build context
        context = ValidationContext(
            nf_found=nf_found,
            valor_pago=valor_pago,
            valor_documento=valor_documento,
            valor_extracted=valor_extracted,
            tipo_documento=tipo_documento,
            page_categories=page_categories,
            date_valid=date_valid,
            is_duplicate=is_duplicate,
            data_emissao_expected=data_emissao_expected,
            data_emissao_extracted=data_emissao_extracted,
            data_servico=data_servico,
            cnpj_data_abertura=cnpj_data_abertura
        )

        # Evaluate rules
        result = self.rule_engine.evaluate(context)

        return result  # Return full RuleResult with classification, rule_name, reason

    def _is_duplicate_nf(
        self,
        cnpj_norm: str,
        numero_norm: str,
        cod_organizacao: str,
        cod_unidade: str,
        pdf_name: str,
        id_documento: int | None = None,
        data_envio: str | None = None,
    ) -> bool:
        """
        Check if NF is a duplicate using 4-field logic.

        Duplicate Detection Rules (from TODO in duplicate_nf.py):
        1. Dedup key: (cnpj, numero_nf, cod_organizacao, cod_unidade) - exact 4-tuple must repeat
        2. Special case: If cod_organizacao == cod_unidade (sede/headquarters) → NOT duplicate
        3. Order by (data_envio, id_documento) - FIRST submission is original
        4. All submissions AFTER the first are duplicates

        :param cnpj_norm: Normalized CNPJ.
        :param numero_norm: Normalized numero_nf.
        :param cod_organizacao: Organization code.
        :param cod_unidade: Unit code.
        :param pdf_name: Current PDF name.
        :param id_documento: Document ID (tie-breaker when dates are equal).
        :param data_envio: Submission date.
        :returns: True if this NF is a duplicate (same 4-tuple submitted after
            the first), False if this is the first submission or is "sede"
            (headquarters).
        """
        # Special case: "sede" (headquarters) - never duplicate
        # When cod_organizacao == cod_unidade, it's a headquarters unit with special business rules
        if cod_organizacao and cod_unidade and cod_organizacao == cod_unidade:
            return False

        # Build 4-field dedup key: exact combination must repeat
        cod_org_norm = cod_organizacao if cod_organizacao else ''
        cod_unit_norm = cod_unidade if cod_unidade else ''
        dedup_key = (cnpj_norm, numero_norm, cod_org_norm, cod_unit_norm)

        pdf_list = self.deduplication_lookup.get(dedup_key, [])

        # If this exact 4-tuple appears in only 1 PDF or less → not a duplicate
        if len(pdf_list) <= 1:
            return False

        # Get current submission date
        if data_envio is None:
            data_envio = self.pdf_data_envio.get(pdf_name)

        if not data_envio:
            # No date info → fall back to simple count-based logic
            # If 4-tuple appears multiple times, mark all except one as duplicate
            return len(pdf_list) > 1

        # Check if any other record with same 4-tuple has EARLIER (data_envio, id_documento)
        from datetime import datetime

        for entry in pdf_list:
            other_pdf = entry['pdf_name']
            other_date = entry.get('data_envio')
            other_id = entry.get('id_documento')

            # Skip if same PDF
            if other_pdf == pdf_name:
                continue

            # Skip if other date is missing
            if not other_date:
                continue

            # Compare (data_envio, id_documento) tuples
            try:
                # Parse dates for comparison
                if isinstance(data_envio, str):
                    current_date_obj = datetime.fromisoformat(data_envio.replace('/', '-'))
                else:
                    current_date_obj = data_envio

                if isinstance(other_date, str):
                    other_date_obj = datetime.fromisoformat(other_date.replace('/', '-'))
                else:
                    other_date_obj = other_date

                # Compare dates first
                if other_date_obj < current_date_obj:
                    return True  # Other is earlier → current is duplicate
                elif other_date_obj == current_date_obj:
                    # Dates are equal → use id_documento as tie-breaker
                    if id_documento is not None and other_id is not None:
                        if other_id < id_documento:
                            return True  # Other has lower ID → current is duplicate

            except Exception:
                # If date parsing fails, fall back to simple comparison
                try:
                    # Compare as strings/integers
                    if other_date < data_envio:
                        return True
                    elif other_date == data_envio and id_documento is not None and other_id is not None:
                        if other_id < id_documento:
                            return True
                except Exception:
                    pass

        # No earlier submission found → current is the first (original)
        return False

    def _find_standard_matches(
        self,
        cnpj_declarado: str,
        numero_declarado: str,
        data_declarada: str | None,
        extracted_nfs: list[dict],
        min_match_score: int = 2,
    ) -> list[dict]:
        """
        Find documents that match using 3-field scoring (CNPJ + número + data_emissão).

        Uses flexible matching where any 2 of 3 fields matching is sufficient.
        This handles cases where one field might have minor discrepancies.

        :param cnpj_declarado: Normalized CNPJ from declaration.
        :param numero_declarado: Normalized número from declaration.
        :param data_declarada: Data de emissão from declaration (optional).
        :param extracted_nfs: List of all extracted documents.
        :param min_match_score: Minimum number of fields that must match
            (default: 2).
        :returns: List of documents that match (annotated with _match_score).
        """
        from .utils import DocumentFields, match_score_3_fields

        matches = []

        for ext_nf in extracted_nfs:
            # Calculate match score using all 3 fields
            score = match_score_3_fields(
                expected=DocumentFields(
                    cnpj=cnpj_declarado,
                    numero=numero_declarado,
                    data=data_declarada,
                ),
                extracted=DocumentFields(
                    cnpj=ext_nf.get('cnpj_emitente', ''),
                    numero=ext_nf.get('numero_nf', ''),
                    data=ext_nf.get('data_emissao'),
                ),
            )

            # Only include if score meets minimum threshold
            if score >= min_match_score:
                # Annotate the match with score information
                match_copy = ext_nf.copy()
                match_copy['_match_score'] = score
                matches.append(match_copy)

        return matches

    def _handle_single_match(
        self,
        expected_nf: dict,
        extracted: dict,
        cnpj_start_dates: dict,
        page_categories: list[str],
    ) -> dict:
        """
        Handle case with single document match.

        :param expected_nf: Expected declaration dict.
        :param extracted: Single matched document.
        :param cnpj_start_dates: CNPJ → start_date cache.
        :param page_categories: Page categories list.
        :returns: Result dict with extracted document and classification.
        """
        ext_cnpj_norm = normalize_cnpj(extracted.get('cnpj_emitente', ''))
        ext_numero_norm = normalize_number(extracted.get('numero_nf', ''))
        ext_valor_norm = normalize_value(extracted.get('valor_total', 0.0))

        # Check if valor matches
        valor_match = abs(expected_nf['valor_norm'] - ext_valor_norm) < 0.01

        # Determine match type
        numero_match_type = 'exact' if expected_nf['numero_norm'] == ext_numero_norm else 'fuzzy'

        # Validate date against company start date
        data_envio = expected_nf['original'].get('data_envio')
        inicio_atividade = expected_nf['original'].get('cnpj_data_abertura') or cnpj_start_dates.get(ext_cnpj_norm)
        date_valid = check_date_against_company_start(data_envio, inicio_atividade)

        # Check for deduplication
        is_duplicate = self._is_duplicate_nf(
            ext_cnpj_norm,
            ext_numero_norm,
            expected_nf.get('cod_organizacao', ''),
            expected_nf.get('cod_unidade', ''),
            expected_nf.get('pdf_name', ''),
            expected_nf.get('id_documento'),
            expected_nf['original'].get('data_envio')
        )

        # Compute classification
        rule_result = self._classify(
            nf_found=True,
            valor_pago=expected_nf['original'].get('valor_pago'),
            valor_documento=expected_nf['valor_norm'],
            valor_extracted=ext_valor_norm,
            tipo_documento=extracted.get('tipo_documento'),
            page_categories=page_categories,
            date_valid=date_valid,
            is_duplicate=is_duplicate,
            data_emissao_expected=expected_nf['original'].get('data_emissao'),
            data_emissao_extracted=extracted.get('data_emissao'),
            data_servico=extracted.get('data_servico'),
            cnpj_data_abertura=inicio_atividade
        )

        return {
            'extracted': extracted,
            'expected': expected_nf['original'],
            'valor_match': valor_match,
            'match_quality': 'PERFECT' if valor_match else 'GOOD',
            'classification': rule_result.classification,
            'rule_name': rule_result.rule_name,
            'reason': rule_result.reason,
            'numero_match_type': numero_match_type
        }

    def _handle_multiple_matches(
        self,
        expected_nf: dict,
        matches: list[dict],
        cnpj_start_dates: dict,
        page_categories: list[str],
    ) -> dict:
        """
        Handle case with multiple document matches (prioritization).

        :param expected_nf: Expected declaration dict.
        :param matches: List of matched documents.
        :param cnpj_start_dates: CNPJ → start_date cache.
        :param page_categories: Page categories list.
        :returns: Result dict with prioritized document and classification.
        """
        from .document_prioritizer import select_prioritized_document

        # Log warning about multiple matches
        types = [doc.get('tipo_documento') for doc in matches]
        logger.warning(
            f"Multiple documents match CNPJ {expected_nf['cnpj_norm']} + "
            f"Número {expected_nf['numero_norm']}: {types}. "
            f"Selecting by priority."
        )

        # Select prioritized document
        best_extracted = select_prioritized_document(matches)

        # Process as single match
        return self._handle_single_match(
            expected_nf,
            best_extracted,
            cnpj_start_dates,
            page_categories
        )

    def _handle_nf_ticket_merge(
        self,
        expected_nf: dict,
        nf: dict,
        ticket: dict,
        match_type: str,
        cnpj_start_dates: dict,
        page_categories: list[str],
    ) -> dict:
        """
        Handle NF + Ticket merge case.

        :param expected_nf: Expected declaration dict.
        :param nf: NF document.
        :param ticket: Ticket document.
        :param match_type: 'direct' or 'reverse'.
        :param cnpj_start_dates: CNPJ → start_date cache.
        :param page_categories: Page categories list.
        :returns: Result dict with merged document and classification.
        """
        from .document_merger import get_merge_justificativa, merge_nf_and_ticket
        from .rps_matcher import get_apontamento_leve_justification, should_apply_apontamento_leve

        # Merge documents
        merged = merge_nf_and_ticket(nf, ticket)

        # Generate justification
        justificativa = get_merge_justificativa(nf, ticket, merged)

        # Check if Apontamento Leve should be applied
        numero_declarado = expected_nf['original'].get('numero_nf', '')
        if should_apply_apontamento_leve(numero_declarado, nf, ticket, match_type):
            classificacao_especial = 'Apontamento Leve'
            justificativa = get_apontamento_leve_justification(nf, ticket, numero_declarado)
        else:
            classificacao_especial = None

        # Process merged document as single match
        result = self._handle_single_match(
            expected_nf,
            merged,
            cnpj_start_dates,
            page_categories
        )

        # Override justification and classification if applicable
        result['merge_justificativa'] = justificativa
        if classificacao_especial:
            result['classification'] = classificacao_especial
            result['reason'] = justificativa

        return result

    def _handle_missing_nf(
        self,
        expected_nf: dict,
        page_categories: list[str],
    ) -> dict:
        """
        Handle case with no document match (missing NF).

        :param expected_nf: Expected declaration dict.
        :param page_categories: Page categories list.
        :returns: Result dict for missing NF.
        """
        # Check for deduplication
        is_duplicate = self._is_duplicate_nf(
            expected_nf['cnpj_norm'],
            expected_nf['numero_norm'],
            expected_nf.get('cod_organizacao', ''),
            expected_nf.get('cod_unidade', ''),
            expected_nf.get('pdf_name', ''),
            expected_nf.get('id_documento'),
            expected_nf['original'].get('data_envio')
        )

        # Compute classification for missing NF
        rule_result = self._classify(
            nf_found=False,
            valor_pago=expected_nf['original'].get('valor_pago'),
            valor_documento=expected_nf['valor_norm'],
            valor_extracted=None,
            tipo_documento=None,
            page_categories=page_categories,
            is_duplicate=is_duplicate,
            data_emissao_expected=expected_nf['original'].get('data_emissao'),
            data_emissao_extracted=None
        )

        missing_nf_data = expected_nf['original'].copy()
        missing_nf_data['classification'] = rule_result.classification
        missing_nf_data['rule_name'] = rule_result.rule_name
        missing_nf_data['reason'] = rule_result.reason

        return {
            'expected': missing_nf_data,
            'extracted': None,
            'classification': rule_result.classification,
            'rule_name': rule_result.rule_name,
            'reason': rule_result.reason
        }

    def _process_single_declaration(
        self,
        expected_nf: dict,
        extracted_nfs: list[dict],
        cnpj_start_dates: dict,
        page_categories: list[str],
    ) -> dict:
        """
        Process a SINGLE declaration independently (declaração-centric).

        This is the core of the new strategy:
        1. Try RPS match first (NF + Ticket)
        2. If no RPS match, try standard match (CNPJ + número)
        3. Handle based on number of matches

        :param expected_nf: Expected declaration dict (already normalized).
        :param extracted_nfs: List of ALL extracted documents from PDF.
        :param cnpj_start_dates: CNPJ → start_date cache.
        :param page_categories: Page categories list.
        :returns: Result dict for this declaration.
        """
        from .rps_matcher import find_nf_ticket_by_rps

        # STEP 1: Try RPS match (NF + Ticket)
        nf, ticket, match_type = find_nf_ticket_by_rps(
            expected_nf['original'],
            extracted_nfs
        )

        if nf and ticket:
            # Case C: NF + Ticket merge
            logger.info(
                f"NF+Ticket merge detected for {expected_nf['cnpj_norm']} / "
                f"{expected_nf['numero_norm']} (match_type: {match_type})"
            )
            return self._handle_nf_ticket_merge(
                expected_nf,
                nf,
                ticket,
                match_type,
                cnpj_start_dates,
                page_categories
            )

        # STEP 2: Standard match (CNPJ + número + data)
        # Strategy: always try perfect matches (3/3) first.
        # If min_match_score < 3, fall back to partial matches when no perfect match found.
        # When min_match_score == 3, the fallback is skipped entirely.

        # First, try to find perfect matches (score = 3)
        perfect_matches = self._find_standard_matches(
            expected_nf['cnpj_norm'],
            expected_nf['numero_norm'],
            expected_nf['original'].get('data_emissao'),
            extracted_nfs,
            min_match_score=3  # Require all 3 fields to match
        )

        # If no perfect match found, fall back to partial matches only when configured to do so
        if len(perfect_matches) == 0 and self.min_match_score < 3:
            standard_matches = self._find_standard_matches(
                expected_nf['cnpj_norm'],
                expected_nf['numero_norm'],
                expected_nf['original'].get('data_emissao'),
                extracted_nfs,
                min_match_score=self.min_match_score  # Allow partial match per config
            )
        else:
            # Use perfect matches (or empty list when min_match_score == 3 and no 3/3 found)
            standard_matches = perfect_matches

        # STEP 3: Handle based on number of matches
        if len(standard_matches) == 0:
            # No match
            return self._handle_missing_nf(expected_nf, page_categories)

        elif len(standard_matches) == 1:
            # Case A: Single match
            return self._handle_single_match(
                expected_nf,
                standard_matches[0],
                cnpj_start_dates,
                page_categories
            )

        else:
            # Case B: Multiple matches (prioritization)
            return self._handle_multiple_matches(
                expected_nf,
                standard_matches,
                cnpj_start_dates,
                page_categories
            )

    def validate_extraction(
        self,
        pdf_name: str,
        extracted_nfs: list[dict],
        page_categories: list[str] | None = None,
    ) -> dict:
        """
        Validate extraction results for a single PDF against expected NFs.

        :param pdf_name: Name of PDF file.
        :param extracted_nfs: List of extracted NF dicts with keys:
            - cnpj_emitente: Extracted CNPJ
            - numero_nf: Extracted NF number
            - valor_total or valor_total_servico: Extracted value
            - tipo_documento: Document type (optional, for classification)
        :param page_categories: List of page categories from classifier
            (optional, for classification).
        :returns: Validation result dict with:
            - status: 'OK', 'WARNINGS', or 'PROBLEMS'
            - correctly_extracted: List of correctly matched NFs (with classification)
            - missing_nfs: List of expected NFs not found (with classification)
            - suspicious_extractions: List of extracted NFs not in expected list
            - normalization_issues: List of NFs with normalization mismatches
            - summary: Dict with counts (including classification breakdown)
        """
        # Get expected NFs for this PDF
        expected_list = self.expected_by_pdf.get(pdf_name, [])

        # TODO: move the composal of this object to a separate function
        # Build CNPJ → inicio_atividade_data cache from BigQuery
        cnpj_start_dates = {}

        # Get all unique CNPJs from extracted NFs
        unique_cnpjs = set(
            normalize_cnpj(nf.get('cnpj_emitente', ''))
            for nf in extracted_nfs
        )

        # Query BigQuery for each unique CNPJ
        from run_poc.bigquery_loader import (
            get_company_start_date,
        )

        for cnpj in unique_cnpjs:
            if cnpj:  # Skip empty CNPJs
                try:
                    start_date = get_company_start_date(cnpj)
                    cnpj_start_dates[cnpj] = start_date
                except Exception as e:
                    # Log error but continue (validation will skip if start date is missing)
                    logger.warning("Failed to query company start date for %s: %s", cnpj, e)

        # If no expected NFs for this PDF, all extractions are suspicious
        if not expected_list:
            return {
                'status': 'WARNINGS' if len(extracted_nfs) > 0 else 'OK',
                'correctly_extracted': [],
                'missing_nfs': [],
                'suspicious_extractions': extracted_nfs,
                'normalization_issues': [],
                'summary': {
                    'total_expected': 0,
                    'total_extracted': len(extracted_nfs),
                    'correctly_extracted': 0,
                    'missing': 0,
                    'suspicious': len(extracted_nfs),
                    'normalization_issues': 0
                },
                'message': f'No expected NFs for {pdf_name} - all {len(extracted_nfs)} extractions are suspicious'
            }

        # NEW STRATEGY: Process each declaration independently (declaração-centric)
        logger.info(f"Processing {len(expected_list)} declarations for {pdf_name}")

        correctly_extracted = []
        missing_nfs = []

        for exp_nf in expected_list:
            # Store pdf_name in expected_nf for deduplication check
            exp_nf['pdf_name'] = pdf_name

            # Process this single declaration
            result = self._process_single_declaration(
                exp_nf,
                extracted_nfs,
                cnpj_start_dates,
                page_categories or []
            )

            # Categorize result
            if result.get('extracted'):
                # Match found
                correctly_extracted.append(result)
            else:
                # Missing NF
                missing_nfs.append(result['expected'])

        # Determine overall status
        has_problems = len(missing_nfs) > 0

        if has_problems:
            status = 'PROBLEMS'
        else:
            status = 'OK'

        # Count classifications
        classification_counts = {'OK': 0, 'Suspect': 0, 'Not Analyzable': 0, 'Apontamento Leve': 0}

        # Count from correctly extracted
        for item in correctly_extracted:
            classification_counts[item.get('classification', 'OK')] += 1

        # Count from missing
        for item in missing_nfs:
            classification_counts[item.get('classification', 'Suspect')] += 1

        logger.info(
            f"Validation complete for {pdf_name}: "
            f"{len(correctly_extracted)} matched, {len(missing_nfs)} missing"
        )

        return {
            'status': status,
            'correctly_extracted': correctly_extracted,
            'missing_nfs': missing_nfs,
            'suspicious_extractions': [],  # Not used in new strategy
            'normalization_issues': [],    # Not used in new strategy
            'summary': {
                'total_expected': len(expected_list),
                'total_extracted': len(extracted_nfs),
                'correctly_extracted': len(correctly_extracted),
                'missing': len(missing_nfs),
                'suspicious': 0,  # Not used in new strategy
                'normalization_issues': 0,  # Not used in new strategy
                # Classification breakdown
                'classification_ok': classification_counts['OK'],
                'classification_suspect': classification_counts['Suspect'],
                'classification_not_analyzable': classification_counts['Not Analyzable']
            }
        }

    def validate_batch(self, extraction_results: list[dict]) -> dict:
        """
        Validate a batch of extraction results.

        :param extraction_results: List of dicts with:
            - pdf_name: Name of PDF
            - extracted_nfs: List of extracted NFs
        :returns: Batch validation result with aggregated statistics.
        """
        validations = []

        for result in extraction_results:
            pdf_name = result['pdf_name']
            extracted_nfs = result.get('extracted_nfs', [])

            validation = self.validate_extraction(pdf_name, extracted_nfs)
            validation['pdf_name'] = pdf_name
            validations.append(validation)

        # Aggregate statistics
        total_expected = sum(v['summary']['total_expected'] for v in validations)
        total_extracted = sum(v['summary']['total_extracted'] for v in validations)
        total_correct = sum(v['summary']['correctly_extracted'] for v in validations)
        total_missing = sum(v['summary']['missing'] for v in validations)
        total_suspicious = sum(v['summary']['suspicious'] for v in validations)
        total_norm_issues = sum(v['summary']['normalization_issues'] for v in validations)

        # Aggregate classification counts
        total_ok = sum(v['summary'].get('classification_ok', 0) for v in validations)
        total_suspect = sum(v['summary'].get('classification_suspect', 0) for v in validations)
        total_not_analyzable = sum(v['summary'].get('classification_not_analyzable', 0) for v in validations)

        pdfs_with_problems = sum(1 for v in validations if v['status'] == 'PROBLEMS')
        pdfs_with_warnings = sum(1 for v in validations if v['status'] == 'WARNINGS')
        pdfs_ok = sum(1 for v in validations if v['status'] == 'OK')

        # Calculate metrics
        precision = 100 * total_correct / total_extracted if total_extracted > 0 else 0
        recall = 100 * total_correct / total_expected if total_expected > 0 else 0

        return {
            'validations': validations,
            'aggregate_summary': {
                'total_pdfs': len(validations),
                'pdfs_with_problems': pdfs_with_problems,
                'pdfs_with_warnings': pdfs_with_warnings,
                'pdfs_ok': pdfs_ok,
                'total_expected_nfs': total_expected,
                'total_extracted_nfs': total_extracted,
                'correctly_extracted': total_correct,
                'missing_nfs': total_missing,
                'suspicious_extractions': total_suspicious,
                'normalization_issues': total_norm_issues,
                'precision': precision,
                'recall': recall,
                # Classification breakdown
                'classification_ok': total_ok,
                'classification_suspect': total_suspect,
                'classification_not_analyzable': total_not_analyzable
            }
        }

    def print_validation_report(self, validation: dict, verbose: bool = True):
        """
        Log a human-readable validation report for a single PDF.

        :param validation: Validation result dict.
        :param verbose: If True, show detailed lists of NFs.
        """
        pdf_name = validation.get('pdf_name', 'Unknown')
        status = validation['status']
        summary = validation['summary']

        status_symbol = {'OK': '[OK]', 'WARNINGS': '[WARNING]', 'PROBLEMS': '[PROBLEM]'}.get(status, '[?]')

        logger.info("%s %s - Status: %s", status_symbol, pdf_name, status)
        logger.info("  Expected: %d NFs", summary['total_expected'])
        logger.info("  Extracted: %d NFs", summary['total_extracted'])
        logger.info("  Correctly extracted: %d", summary['correctly_extracted'])

        if summary['missing'] > 0:
            logger.warning("  [!] Missing: %d NFs", summary['missing'])

        if summary['suspicious'] > 0:
            logger.warning("  [!] Suspicious: %d extractions", summary['suspicious'])

        if summary['normalization_issues'] > 0:
            logger.warning("  [!] Normalization issues: %d", summary['normalization_issues'])

        if verbose and status != 'OK':
            if validation['missing_nfs']:
                logger.info("  Missing NFs:")
                for nf in validation['missing_nfs']:
                    logger.info(
                        "    - CNPJ: %s, Numero: %s, Valor: %s",
                        nf['cnpj'], nf['numero_nf'], nf['valor_total'],
                    )

            if validation['suspicious_extractions']:
                logger.info("  Suspicious Extractions:")
                for item in validation['suspicious_extractions']:
                    ext = item['extracted']
                    logger.info(
                        "    - CNPJ: %s, Numero: %s | Reason: %s",
                        ext.get('cnpj_emitente', 'N/A'), ext.get('numero_nf', 'N/A'), item['reason'],
                    )

            if validation['normalization_issues']:
                logger.info("  Normalization Issues:")
                for item in validation['normalization_issues']:
                    exp = item['expected']
                    logger.info(
                        "    - Expected: %s | Extracted: %s | Issue: %s",
                        exp['numero_nf'], item['extracted_numero'], item['issue'],
                    )

    def print_batch_report(self, batch_validation: dict, group_by_nf: bool = True):
        """
        Log aggregate batch validation report.

        :param batch_validation: Batch validation result dict.
        :param group_by_nf: If True, group results by searched NF instead of by PDF.
        """
        summary = batch_validation['aggregate_summary']

        sep = "=" * 80
        logger.info(sep)
        logger.info("COMPLIANCE VALIDATION REPORT")
        logger.info(sep)
        logger.info("Total PDFs: %d", summary['total_pdfs'])
        logger.info("  [OK] Status OK: %d", summary['pdfs_ok'])
        logger.info("  [WARN] Warnings: %d", summary['pdfs_with_warnings'])
        logger.info("  [PROB] Problems: %d", summary['pdfs_with_problems'])
        logger.info("Total Expected NFs: %d", summary['total_expected_nfs'])
        logger.info("Total Extracted NFs: %d", summary['total_extracted_nfs'])
        logger.info("Correctly Extracted: %d", summary['correctly_extracted'])
        logger.info("Missing NFs: %d", summary['missing_nfs'])
        logger.info("Suspicious Extractions: %d", summary['suspicious_extractions'])
        logger.info("Normalization Issues: %d", summary['normalization_issues'])
        logger.info("Classification Breakdown:")
        logger.info("  [OK] OK: %d", summary.get('classification_ok', 0))
        logger.info("  [SUSPECT] Suspect: %d", summary.get('classification_suspect', 0))
        logger.info("  [N/A] Not Analyzable: %d", summary.get('classification_not_analyzable', 0))
        logger.info("Precision: %.2f%%", summary['precision'])
        logger.info("Recall: %.2f%%", summary['recall'])
        logger.info(sep)

        if group_by_nf:
            self._print_nf_centric_report(batch_validation)

    def _print_nf_centric_report(self, batch_validation: dict):
        """
        Print NF-centric view: show status of each searched NF.

        :param batch_validation: Batch validation result dict.
        """
        sep = "=" * 80
        logger.info(sep)
        logger.info("SEARCHED NFs - DETAILED STATUS")
        logger.info(sep)

        # Collect all NF statuses
        nf_statuses = []

        for validation in batch_validation['validations']:
            pdf_name = validation['pdf_name']

            for nf_data in validation['correctly_extracted']:
                expected = nf_data['expected']
                extracted = nf_data['extracted']
                nf_statuses.append({
                    'pdf_name': pdf_name,
                    'cnpj': expected['cnpj'],
                    'numero_nf': expected['numero_nf'],
                    'valor_total': expected['valor_total'],
                    'page': expected.get('page', 'Unknown'),
                    'status': 'FOUND',
                    'match_quality': nf_data['match_quality'],
                    'extracted_valor': extracted.get('valor_total', 0.0),
                    'classification': nf_data.get('classification', 'Unknown'),
                })

            for nf_data in validation['missing_nfs']:
                nf_statuses.append({
                    'pdf_name': pdf_name,
                    'cnpj': nf_data['cnpj'],
                    'numero_nf': nf_data['numero_nf'],
                    'valor_total': nf_data['valor_total'],
                    'page': nf_data.get('page', 'Unknown'),
                    'status': 'MISSING',
                    'match_quality': None,
                    'extracted_valor': None,
                    'classification': nf_data.get('classification', 'Unknown'),
                })

            for nf_data in validation['normalization_issues']:
                expected = nf_data['expected']
                nf_statuses.append({
                    'pdf_name': pdf_name,
                    'cnpj': expected['cnpj'],
                    'numero_nf': expected['numero_nf'],
                    'valor_total': expected['valor_total'],
                    'page': expected.get('page', 'Unknown'),
                    'status': 'NORM_ISSUE',
                    'match_quality': None,
                    'extracted_numero': nf_data['extracted_numero'],
                    'issue': nf_data['issue'],
                })

        status_order = {'FOUND': 0, 'NORM_ISSUE': 1, 'MISSING': 2}
        nf_statuses.sort(key=lambda x: (status_order.get(x['status'], 3), x['pdf_name'], x['cnpj']))

        found_count = 0
        missing_count = 0
        norm_issue_count = 0

        for nf in nf_statuses:
            if nf['status'] == 'FOUND':
                found_count += 1
                logger.info(
                    "[OK] FOUND | PDF: %s | CNPJ: %s | Numero: %s | "
                    "Valor esperado: %s | Valor extraído: %s | "
                    "Match: %s | Classificação: %s | Página: %s",
                    nf['pdf_name'], nf['cnpj'], nf['numero_nf'],
                    nf['valor_total'], nf['extracted_valor'],
                    nf['match_quality'], nf.get('classification', 'Unknown'), nf['page'],
                )

            elif nf['status'] == 'NORM_ISSUE':
                norm_issue_count += 1
                logger.warning(
                    "[WARN] NORMALIZATION ISSUE | PDF: %s | CNPJ: %s | "
                    "Número esperado: %s | Número extraído: %s | "
                    "Valor: %s | Problema: %s | Página: %s",
                    nf['pdf_name'], nf['cnpj'], nf['numero_nf'],
                    nf['extracted_numero'], nf['valor_total'], nf['issue'], nf['page'],
                )

            elif nf['status'] == 'MISSING':
                missing_count += 1
                logger.warning(
                    "[MISS] NOT FOUND | PDF: %s | CNPJ: %s | Numero: %s | "
                    "Valor: %s | Classificação: %s | Página: %s",
                    nf['pdf_name'], nf['cnpj'], nf['numero_nf'],
                    nf['valor_total'], nf.get('classification', 'Unknown'), nf['page'],
                )

        logger.info(sep)
        logger.info("SUSPICIOUS EXTRACTIONS (Not in expected list)")
        logger.info(sep)

        suspicious_count = 0
        for validation in batch_validation['validations']:
            pdf_name = validation['pdf_name']
            for suspicious in validation['suspicious_extractions']:
                suspicious_count += 1
                extracted = suspicious['extracted']
                logger.warning(
                    "[!] SUSPICIOUS - %s | CNPJ: %s | Numero: %s | Valor: %s | Reason: %s",
                    pdf_name,
                    extracted.get('cnpj_emitente', 'N/A'),
                    extracted.get('numero_nf', 'N/A'),
                    extracted.get('valor_total', 'N/A'),
                    suspicious['reason'],
                )

        if suspicious_count == 0:
            logger.info("No suspicious extractions found.")

        logger.info(sep)
        logger.info("SUMMARY BY SEARCHED NF")
        logger.info(sep)
        logger.info("Found: %d", found_count)
        logger.info("Missing: %d", missing_count)
        logger.info("Normalization Issues: %d", norm_issue_count)
        logger.info("Suspicious Extractions: %d", suspicious_count)
        logger.info(sep)

    @staticmethod
    def load_expected_nfs_from_excel(excel_path: Path) -> list[dict]:
        """
        Load expected NFs from validation Excel file.

        :param excel_path: Path to validation Excel file with NF_Details sheet.
        :returns: List of expected NF dicts.
        :raises KeyError: If the sheet ``NF_Details`` is not found in the workbook.
        :raises ValueError: If required columns are missing from the sheet header.
        """
        from openpyxl import load_workbook  # noqa: PLC0415

        wb = load_workbook(excel_path, read_only=True, data_only=True)
        if 'NF_Details' not in wb.sheetnames:
            raise KeyError(
                f"Sheet 'NF_Details' not found in '{excel_path}'. "
                f"Available sheets: {wb.sheetnames}"
            )
        ws = wb['NF_Details']

        rows = ws.iter_rows(values_only=True)
        header = next(rows)
        col_idx = {name: i for i, name in enumerate(header) if name is not None}

        required_cols = {'PDF_name', 'CNPJ', 'Numero_NF', 'Valor_Total'}
        missing = required_cols - col_idx.keys()
        if missing:
            raise ValueError(
                f"Required columns missing from 'NF_Details' sheet: {sorted(missing)}"
            )

        page_col = col_idx.get('NF_Page')
        expected_nfs = []
        for row in rows:
            expected_nfs.append({
                'pdf_name': row[col_idx['PDF_name']],
                'cnpj': row[col_idx['CNPJ']],
                'numero_nf': row[col_idx['Numero_NF']],
                'valor_total': row[col_idx['Valor_Total']],
                'page': row[page_col] if page_col is not None else 'Unknown',
            })

        wb.close()
        return expected_nfs


# Convenience function for quick validation
def validate_against_expected(
    pdf_name: str,
    extracted_nfs: list[dict],
    expected_nfs: list[dict],
    page_categories: list[str] | None = None,
) -> dict:
    """
    Quick validation of extraction results against expected NFs.

    :param pdf_name: Name of PDF file.
    :param extracted_nfs: List of extracted NF dicts.
    :param expected_nfs: List of expected NF dicts (all PDFs).
    :param page_categories: List of page categories from classifier (optional).
    :returns: Validation result dict with classification
        (OK/Suspect/Not Analyzable).
    """
    validator = ComplianceValidator(expected_nfs)
    return validator.validate_extraction(pdf_name, extracted_nfs, page_categories)
